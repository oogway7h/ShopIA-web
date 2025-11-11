import os
import traceback
from io import BytesIO
from datetime import datetime, timedelta

from django.db.models import Count, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response


from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


try:
    from apps.usuarios.models import Usuario, Rol
except ImportError:
    print("Error: No se pudo importar Usuario/Rol. Usando Mock.")
    class Usuario: objects = type('obj', (object,), {
        'select_related': lambda *a, **k: Usuario.objects,
        'get': lambda *a, **k: None,
        'filter': lambda *a, **k: Usuario.objects,
        'none': lambda *a, **k: [],
        'all': lambda *a, **k: [],
    })()
    class Rol: objects = type('obj', (object,), {
        'get': lambda *a, **k: None,
    })()

# Asumimos que tienes una app 'ventas' con los modelos 'Venta' y 'DetalleVenta'
try:
    from apps.ventas.models import Venta
except ImportError:
    print("Error: No se pudo importar el modelo Venta. Usando Mock.")
    class Venta: objects = type('obj', (object,), {
        'select_related': lambda *a, **k: Venta.objects,
        'all': lambda *a, **k: [],
        'filter': lambda *a, **k: Venta.objects,
        'none': lambda *a, **k: [],
        'annotate': lambda *a, **k: Venta.objects,
        'values': lambda *a, **k: Venta.objects,
        'order_by': lambda *a, **k: [],
    })()

try:
    from apps.ventas.models import DetalleVenta
except ImportError:
    print("Error: No se pudo importar el modelo DetalleVenta. Usando Mock.")
    class DetalleVenta: objects = type('obj', (object,), {
        'select_related': lambda *a, **k: DetalleVenta.objects,
        'all': lambda *a, **k: [],
        'filter': lambda *a, **k: [],
        'none': lambda *a, **k: [],
    })()
    


try:
    from apps.productos.models import Categoria, Producto
except ImportError:
    print("Error: No se pudo importar Usuario/Rol. Usando Mock.")

try:
    from .nlp_service import procesar_comando_voz
except ImportError:
    print("ADVERTENCIA: No se pudo importar 'nlp_service'.")
    def procesar_comando_voz(texto):
        return {"error": "Servicio NLP no cargado."}


FONT_NAME = 'Helvetica'
FONT_BOLD = 'Helvetica-Bold'

def _get_optional_date_range(request):
    """
    Obtiene un rango de fechas (fecha_inicio, fecha_fin) desde los query_params.
    Devuelve None si no se proveen.
    """
    fecha_inicio = None
    fecha_fin = None
    try:
        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)
        
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            
    except (ValueError, TypeError):
        print(f"[WARN] Fechas inválidas recibidas: {fecha_inicio_str}, {fecha_fin_str}")
        pass 
    return fecha_inicio, fecha_fin

def _get_pdf_styles():
    """Devuelve los estilos base de ReportLab para los PDF."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'], alignment=TA_CENTER, fontName=FONT_BOLD)
    normal_center = ParagraphStyle('normal_center', parent=styles['Normal'], alignment=TA_CENTER, fontName=FONT_NAME, fontSize=9)
    header_style = ParagraphStyle('header', parent=normal_center, fontName=FONT_BOLD, textColor=colors.whitesmoke)
    normal_left = ParagraphStyle('normal_left', parent=styles['Normal'], alignment=TA_LEFT, fontName=FONT_NAME)
    normal_right = ParagraphStyle('normal_right', parent=styles['Normal'], alignment=TA_RIGHT, fontName=FONT_NAME, fontSize=9)
    return title_style, header_style, normal_center, normal_left, normal_right

def _get_user_permission(request):
    usuario_perfil = request.user
    
    if not usuario_perfil or not usuario_perfil.is_authenticated:
         raise PermissionError("Usuario no autenticado.")

    is_super_admin = usuario_perfil.is_superuser
    
    is_admin = usuario_perfil.roles.filter(nombre='admin').exists()

    if not (is_super_admin or is_admin):
        raise PermissionError("Acceso denegado. Se requiere rol de Administrador.")
        
    return usuario_perfil


#listado de clientes en pdf
@api_view(['GET']) 
@permission_classes([IsAuthenticated])
def generar_reporte_clientes_pdf(request):
    
    try:
        usuario_perfil = _get_user_permission(request)
    except Usuario.DoesNotExist:
        return HttpResponse(f"Error: Perfil de usuario no encontrado.", status=403)
    except PermissionError as e:
        return HttpResponse(str(e), status=403)
    except Exception as e:
        return HttpResponse(f"Error obteniendo perfil de usuario: {e}", status=500)

    fecha_inicio, fecha_fin = _get_optional_date_range(request)
    
    try:
        # Filtra usuarios que tengan el rol 'cliente'
        cliente_rol = Rol.objects.get(nombre='cliente')
        clientes_qs = Usuario.objects.filter(roles=cliente_rol).order_by('nombre')
        
        if fecha_inicio and fecha_fin:
            # Filtra por 'date_joined' (campo del modelo Usuario)
            clientes_qs = clientes_qs.filter(date_joined__date__range=[fecha_inicio, fecha_fin])
            print(f"Filtrando PDF de Clientes por fechas: {fecha_inicio} a {fecha_fin}")
        
        clientes_filtrados = clientes_qs.all()
        titulo_reporte = "Listado General de Clientes de ShopIA"
        total = clientes_filtrados.count()

    except Rol.DoesNotExist:
         return HttpResponse("Error: El Rol 'cliente' no existe en la base de datos.", status=500)
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la base de datos (ver consola).", status=500)

    # Estilos
    title_style, header_style, normal_center, normal_left, normal_right = _get_pdf_styles()

    # Encabezados de la tabla
    data = [
        [
            Paragraph("<b>ID Cliente</b>", header_style),
            Paragraph("<b>Nombre Completo</b>", header_style),
            Paragraph("<b>Correo Electrónico</b>", header_style),
            Paragraph("<b>Teléfono</b>", header_style),
            Paragraph("<b>Estado</b>", header_style),
        ]
    ]

    # Llenar datos de la tabla
    try:
        for cliente in clientes_filtrados: 
            data.append([
                Paragraph(str(cliente.id), normal_center),
                Paragraph(f"{cliente.nombre} {cliente.apellido}", normal_center),
                Paragraph(str(cliente.correo), normal_center),
                Paragraph(str(cliente.telefono or 'N/A'), normal_center),
                Paragraph("Activo" if cliente.estado else "Inactivo", normal_center),
            ])
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al preparar datos del PDF (ver consola).", status=500)

    if len(data) == 1: 
        data.append([Paragraph("No se encontraron clientes registrados.", normal_center), "", "", "", ""])

    # Construir el PDF
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72, title=titulo_reporte)
        elements = []
        
        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {fecha_gen} por {usuario_perfil.nombre}", normal_right))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph(titulo_reporte, title_style))
        elements.append(Spacer(1, 24))

        intro_texto = f"Este reporte detalla <b>{total} cliente(s)</b>"
        if fecha_inicio and fecha_fin:
            intro_texto += f" registrado(s) entre <b>{fecha_inicio}</b> y <b>{fecha_fin}</b>."
        else:
            intro_texto += " (histórico completo)."
        elements.append(Paragraph(intro_texto, normal_left))
        elements.append(Spacer(1, 24))

        table = Table(data, colWidths=[60, 140, 160, 80, 60], repeatRows=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4A90E2")), # Azul
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6F7FF")), 
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        
        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()

        if not pdf_bytes or not pdf_bytes.startswith(b'%PDF-'):
            return HttpResponse("Error: El contenido generado no es un PDF válido.", status=500)
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="listado_clientes.pdf"'
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el PDF de Clientes.", status=500)


#listado de clientes en excel 
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_clientes_excel(request):
    """
    Genera un listado en Excel de todos los usuarios con el rol 'Cliente'.
    """
    try:
        usuario_perfil = _get_user_permission(request)
    except Exception as e:
        return HttpResponse(f"Error de permisos: {e}", status=403)

    fecha_inicio, fecha_fin = _get_optional_date_range(request)

    try:
        cliente_rol = Rol.objects.get(nombre='cliente')
        clientes_qs = Usuario.objects.filter(roles=cliente_rol).order_by('nombre')
        
        if fecha_inicio and fecha_fin:
            clientes_qs = clientes_qs.filter(date_joined=[fecha_inicio, fecha_fin])
        
        clientes_filtrados = clientes_qs.all()
        titulo_reporte = "Reporte de Clientes"

    except Rol.DoesNotExist:
         return HttpResponse("Error: El Rol 'cliente' no existe.", status=500)
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la BD (ver consola).", status=500)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004A99", end_color="004A99", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ["ID", "Nombre", "Apellido", "Correo", "Teléfono", "Sexo", "Estado", "Fecha Registro"]
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Datos
        row_num = 2
        for cliente in clientes_filtrados:
            fecha_local = timezone.localtime(cliente.date_joined)
            fecha_naive = fecha_local.replace(tzinfo=None)
            ws.append([
                cliente.id,
                cliente.nombre,
                cliente.apellido,
                cliente.correo,
                cliente.telefono,
                cliente.get_sexo_display() if cliente.sexo else 'N/A',
                "Activo" if cliente.estado else "Inactivo",
                fecha_naive
            ])
            ws[f'H{row_num}'].number_format = 'YYYY-MM-DD HH:MM'
            row_num += 1

        # Auto-ajustar columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'reporte_clientes_{fecha_inicio or "inicio"}_a_{fecha_fin or "fin"}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el Excel de Clientes.", status=500)


#reportes de ventas en pdf
@api_view(['GET']) 
@permission_classes([IsAuthenticated])
def generar_reporte_ventas_pdf(request):
    """
    Genera un listado en PDF de todas las Ventas, filtrado por fecha,
    categoría y/o producto.
    """
    try:
        usuario_perfil = _get_user_permission(request)
    except Exception as e:
        return HttpResponse(f"Error de permisos: {e}", status=403)

    fecha_inicio, fecha_fin = _get_optional_date_range(request)
    categoria_id = request.query_params.get('categoria_id', None)
    producto_id = request.query_params.get('producto_id', None)
    
    titulo_reporte = "Reporte General de Ventas"
    filtros_aplicados = []
    
    try:
        ventas_qs = Venta.objects.select_related('usuario').order_by('-fecha')
        
        if fecha_inicio and fecha_fin:
            ventas_qs = ventas_qs.filter(fecha__date__range=[fecha_inicio, fecha_fin])
            filtros_aplicados.append(f"Fechas: {fecha_inicio} a {fecha_fin}")
            print(f"[DEBUG] Filtrando PDF de Ventas por fechas: {fecha_inicio} a {fecha_fin}")
        
        
        filtro_producto_aplicado = False
        if producto_id:
            try:
                ventas_qs = ventas_qs.filter(detalles__producto__id=producto_id)
                filtro_producto_aplicado = True
                
                producto_obj = Producto.objects.get(id=producto_id)
                filtros_aplicados.append(f"Producto: {producto_obj.nombre}")
                
            except Exception as e:
                print(f"[WARN] Error al filtrar por producto_id {producto_id}: {e}")
                pass

        if categoria_id and not filtro_producto_aplicado:
            try:
                ventas_qs = ventas_qs.filter(detalles__producto__categoria__id=categoria_id)
                
                categoria_obj = Categoria.objects.get(id=categoria_id)
                filtros_aplicados.append(f"Categoría: {categoria_obj.nombre}")
                
            except Exception as e:
                print(f"[WARN] Error al filtrar por categoria_id {categoria_id}: {e}")
                pass
        
        if producto_id or categoria_id:
            ventas_qs = ventas_qs.distinct()
            
        if filtros_aplicados:
            titulo_reporte = f"Reporte de Ventas ({', '.join(filtros_aplicados)})"
        
        
        ventas_filtradas = ventas_qs.all()
        total = ventas_filtradas.count()

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la base de datos (Ventas).", status=500)

    title_style, header_style, normal_center, normal_left, normal_right = _get_pdf_styles()

    data = [
        [
            Paragraph("<b>ID Venta</b>", header_style),
            Paragraph("<b>Fecha</b>", header_style),
            Paragraph("<b>Cliente (Correo)</b>", header_style),
            Paragraph("<b>Total (Bs.)</b>", header_style),
            Paragraph("<b>Estado</b>", header_style),
        ]
    ]

    try:
        total_ventas_monto = 0
        for venta in ventas_filtradas: 
            total_ventas_monto += getattr(venta, 'monto_total', 0)
            
            fecha_venta_str = "N/A"
            if hasattr(venta, 'fecha') and venta.fecha:
                fecha_naive = timezone.localtime(venta.fecha)
                fecha_venta_str = fecha_naive.strftime('%Y-%m-%d %H:%M')

            data.append([
                Paragraph(str(getattr(venta, 'id', 'N/A')), normal_center),
                Paragraph(fecha_venta_str, normal_center),
                Paragraph(str(getattr(getattr(venta, 'usuario', None), 'correo', 'N/A')), normal_center),
                Paragraph(f"Bs. {getattr(venta, 'monto_total', 0):.2f}", normal_center),
                Paragraph(str(getattr(venta, 'estado', 'N/A')), normal_center),
            ])
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al preparar datos del PDF (Ventas).", status=500)

    if len(data) == 1: 
        data.append([Paragraph("No se encontraron ventas para este rango.", normal_center), "", "", "", ""])

    # Construir PDF
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72, title=titulo_reporte)
        elements = []
        
        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {fecha_gen} por {usuario_perfil.nombre}", normal_right))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph(titulo_reporte, title_style))
        elements.append(Spacer(1, 24))

        intro_texto = f"Este reporte detalla <b>{total} venta(s)</b>"
        if filtros_aplicados:
             intro_texto += f" aplicando los filtros: <b>{', '.join(filtros_aplicados)}</b>."
        elif fecha_inicio and fecha_fin:
             intro_texto += f" registradas entre <b>{fecha_inicio}</b> y <b>{fecha_fin}</b>."
        else:
            intro_texto += " (histórico completo)."
        
        elements.append(Paragraph(intro_texto, normal_left))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Monto total del periodo: <b>Bs. {total_ventas_monto:.2f}</b>", normal_left))
        elements.append(Spacer(1, 24))

        table = Table(data, colWidths=[60, 100, 160, 80, 80], repeatRows=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#318666")), # Verde
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6F7FF")), 
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()

        if not pdf_bytes or not pdf_bytes.startswith(b'%PDF-'):
            return HttpResponse("Error: El contenido generado no es un PDF válido.", status=500)
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas.pdf"'
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el PDF de Ventas.", status=500)

#reportes de ventas en excel 
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_ventas_excel(request):
    
    try:
        usuario_perfil = _get_user_permission(request)
    except Exception as e:
        return HttpResponse(f"Error de permisos: {e}", status=403)

    fecha_inicio, fecha_fin = _get_optional_date_range(request)
    categoria_id = request.query_params.get('categoria_id', None)
    producto_id = request.query_params.get('producto_id', None)
    
    titulo_reporte = "Reporte General de Ventas"
    filtros_aplicados_str = ""
    
    try:
        ventas_qs = Venta.objects.select_related('usuario').order_by('-fecha')
        
        if fecha_inicio and fecha_fin:
            ventas_qs = ventas_qs.filter(fecha__date__range=[fecha_inicio, fecha_fin])
            filtros_aplicados_str += f"de {fecha_inicio} a {fecha_fin} "

        filtro_producto_aplicado = False
        if producto_id:
            try:
                ventas_qs = ventas_qs.filter(detalles__producto__id=producto_id)
                filtro_producto_aplicado = True
                producto_obj = Producto.objects.get(id=producto_id)
                filtros_aplicados_str += f"Producto: {producto_obj.nombre} "
            except Exception: pass

        if categoria_id and not filtro_producto_aplicado:
            try:
                ventas_qs = ventas_qs.filter(detalles__producto__categoria__id=categoria_id)
                categoria_obj = Categoria.objects.get(id=categoria_id)
                filtros_aplicados_str += f"Categoría: {categoria_obj.nombre} "
            except Exception: pass
        
    
        if producto_id or categoria_id:
            ventas_qs = ventas_qs.distinct()
            
        if filtros_aplicados_str:
            titulo_reporte = f"Reporte de Ventas ({filtros_aplicados_str.strip()})"
        
        
        ventas_filtradas = ventas_qs.all()

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al consultar la BD (Ventas).", status=500)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004A99", end_color="004A99", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        title_cell = ws['F1']
        title_cell.value = titulo_reporte
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center")
        ws.append([]) 
        
        
        headers = ["ID Venta", "Fecha", "Cliente (Nombre)", "Cliente (Correo)", "Total", "Estado"]
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row_num = 4 
        for venta in ventas_filtradas:
            fecha_naive = timezone.localtime(venta.fecha).replace(tzinfo=None)

            ws.append([
                venta.id,
                fecha_naive, 
                f"{venta.usuario.nombre} {venta.usuario.apellido}",
                venta.usuario.correo,
                venta.monto_total, 
                venta.estado,
            ])
            ws[f'B{row_num}'].number_format = 'YYYY-MM-DD HH:MM'
            ws[f'E{row_num}'].number_format = '"Bs." #,##0.00'
            row_num += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'reporte_ventas_{fecha_inicio or "inicio"}_a_{fecha_fin or "fin"}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el Excel de Ventas.", status=500)



#reporte de ventas para hacer los graficos
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_ventas_por_dia_json(request):
    """
    Devuelve un JSON con el total de ventas y número de ventas, agrupado por día.
    Optimizado para gráficos.
    """
    try:
        usuario_perfil= _get_user_permission(request)
    except Exception as e:
        return Response({"error": f"Error de permisos: {e}"}, status=status.HTTP_403_FORBIDDEN)

    # Fechas
    try:
        fecha_fin_dt = timezone.now().date()
        fecha_inicio_dt = fecha_fin_dt - timedelta(days=29) # Últimos 30 días

        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)

        if fecha_inicio_str:
            fecha_inicio_dt = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin_dt = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    
    except ValueError:
        return Response({"error": "Formato de fecha inválido. Use AAAA-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({"error": f"Error procesando fechas: {e}"}, status=status.HTTP_400_BAD_REQUEST)

    # Consulta
    try:
        
        ventas_qs_base = Venta.objects.filter(
            fecha__range=[fecha_inicio_dt, fecha_fin_dt],
            estado='PAGADA'
        )

        data_agrupada = ventas_qs_base.annotate(
            dia=TruncDate('fecha')
        ).values(
            'dia'
        ).annotate(
            total_vendido=Sum('monto_total'), 
            num_ventas=Count('id')      
        ).order_by(
            'dia'
        )
        
        datos_grafico = [
            {
                "fecha": item['dia'].isoformat(), 
                "total_vendido": item['total_vendido'],
                "num_ventas": item['num_ventas']
            } 
            for item in data_agrupada
        ]
        
        
        response_data = {
            "datos_grafico_ventas": datos_grafico,
        }
        
        return Response(response_data, status=status.HTTP_200_OK)

    except Exception as e:
        traceback.print_exc()
        return Response({"error": f"Error al consultar la base de datos: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


#reporte de clientes para hacer graficos
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_clientes_por_mes_json(request):
    """
    Devuelve un JSON con el total de nuevos clientes registrados, agrupado por mes.
    """
    try:
        usuario_perfil = _get_user_permission(request)
    except Exception as e:
        return Response({"error": f"Error de permisos: {e}"}, status=status.HTTP_403_FORBIDDEN)

    # Fechas
    try:
        fecha_fin = timezone.now().date()
        fecha_inicio = fecha_fin - timedelta(days=364) # Último año

        fecha_inicio_str = request.query_params.get('fecha_inicio', None)
        fecha_fin_str = request.query_params.get('fecha_fin', None)

        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
    except Exception as e:
        return Response({"error": f"Formato de fecha inválido: {e}"}, status=status.HTTP_400_BAD_REQUEST)

    # Consulta
    try:
        cliente_rol = Rol.objects.get(nombre='cliente')
        clientes_qs = Usuario.objects.filter(
            roles=cliente_rol,
            date_joined__date__range=[fecha_inicio, fecha_fin]
        )

        
        datos_grafico = (
            clientes_qs
            .annotate(mes_registro=TruncMonth('date_joined'))
            .values('mes_registro')
            .annotate(total=Count('id'))
            .order_by('mes_registro')
        )
        
        datos_grafico_formato = [
            {"mes": item['mes_registro'].strftime('%Y-%m'), "total": item['total']}
            for item in datos_grafico
        ]
        
        return Response({
            "datos_grafico_clientes": datos_grafico_formato,
        }, status=status.HTTP_200_OK)

    except Rol.DoesNotExist:
         return Response({"error": "El Rol 'cliente' no existe."}, status=500)
    except Exception as e:
        traceback.print_exc()
        return Response({"error": "Error al consultar la BD (ver consola)."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



#vista para comando de voz
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def procesar_comando_voz_json(request):
    
    texto_comando = request.data.get('texto_comando', None)
    
    if not texto_comando:
        return Response(
            {"error": "No se proporcionó 'texto_comando'."}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        resultado_nlp = procesar_comando_voz(texto_comando)
        
        if "error" in resultado_nlp:
            return Response(resultado_nlp, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(resultado_nlp, status=status.HTTP_200_OK)
        
    except Exception as e:
        traceback.print_exc()
        return Response(
            {"error": f"Error interno en el servidor NLP: {e}"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    

#views para generar reportes por mes

def get_mas_vendidos_data(request):
    
    fecha_inicio, fecha_fin = _get_optional_date_range(request)
    
    categoria_id = request.query_params.get('categoria', None)
    if not categoria_id:
        categoria_id = request.query_params.get('categoria_id', None)
    
    titulo_reporte = "Reporte de Productos Más Vendidos"
    filtros_aplicados = []

    try:
        detalle_qs = DetalleVenta.objects.all()

        if fecha_inicio and fecha_fin:
            detalle_qs = detalle_qs.filter(venta__fecha__date__range=[fecha_inicio, fecha_fin])
            filtros_aplicados.append(f"Fechas: {fecha_inicio} a {fecha_fin}")
        
        if categoria_id:
            try:
                detalle_qs = detalle_qs.filter(producto__categoria__id=categoria_id)
                categoria_obj = Categoria.objects.get(id=categoria_id)
                filtros_aplicados.append(f"Categoría: {categoria_obj.nombre}")
            except Categoria.DoesNotExist:
                print(f"[WARN] Categoria ID {categoria_id} no encontrada.")
                pass
        
        top_productos_data = detalle_qs.values(
            'producto__id', 
            'producto__nombre',
            'producto__categoria__nombre' 
        ).annotate(
            total_unidades=Sum('cantidad') 
        ).order_by('-total_unidades')[:10] 

        if filtros_aplicados:
            titulo_reporte = f"Top 10 Más Vendidos ({', '.join(filtros_aplicados)})"
        else:
            titulo_reporte = "Top 10 Más Vendidos (Histórico)"

        return top_productos_data, titulo_reporte, fecha_inicio, fecha_fin

    except Exception as e:
        traceback.print_exc()
        raise e 


@api_view(['GET']) 
@permission_classes([IsAuthenticated])
def generar_reporte_mas_vendidos_pdf(request):
    
    try:
        usuario_perfil = _get_user_permission(request)
    except Exception as e:
        return HttpResponse(f"Error de permisos: {e}", status=403)
    
    try:
        top_productos_data, titulo_reporte, fecha_inicio, fecha_fin = get_mas_vendidos_data(request)
    except Exception as e:
        return HttpResponse(f"Error al consultar la base de datos (Más Vendidos): {e}", status=500)

    title_style, header_style, normal_center, normal_left, normal_right = _get_pdf_styles()

    data = [
        [
            Paragraph("<b>Ranking</b>", header_style),
            Paragraph("<b>ID Producto</b>", header_style),
            Paragraph("<b>Nombre Producto</b>", header_style),
            Paragraph("<b>Categoría</b>", header_style),
            Paragraph("<b>Unidades Vendidas</b>", header_style),
        ]
    ]

    try:
        for i, item in enumerate(top_productos_data):
            data.append([
                Paragraph(str(i + 1), normal_center),
                Paragraph(str(item['producto__id']), normal_center),
                Paragraph(str(item['producto__nombre']), normal_center),
                Paragraph(str(item['producto__categoria__nombre']), normal_center),
                Paragraph(str(item['total_unidades']), normal_center),
            ])
    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error al preparar datos del PDF (Más Vendidos).", status=500)

    if len(data) == 1: 
        data.append([Paragraph("No se encontraron productos vendidos para este rango.", normal_center), "", "", "", ""])

    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=72, bottomMargin=72, title=titulo_reporte)
        elements = []
        
        fecha_gen = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        elements.append(Paragraph(f"Generado el: {fecha_gen} por {usuario_perfil.nombre}", normal_right))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(titulo_reporte, title_style))
        elements.append(Spacer(1, 24))

        intro_texto = f"Este reporte detalla los <b>Top {len(top_productos_data)} productos</b> más vendidos"
        if fecha_inicio and fecha_fin:
            intro_texto += f" entre <b>{fecha_inicio}</b> y <b>{fecha_fin}</b>."
        else:
            intro_texto += " histórico completo."
        elements.append(Paragraph(intro_texto, normal_left))
        elements.append(Spacer(1, 24))

        table = Table(data, colWidths=[60, 80, 150, 100, 110], repeatRows=1)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#D6C7DC")), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6F7FF")), 
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ])
        table.setStyle(table_style)
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        buffer.close()
        
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_mas_vendidos.pdf"'
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el PDF de Más Vendidos.", status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_reporte_mas_vendidos_excel(request):
    
    try:
        usuario_perfil = _get_user_permission(request)
    except Exception as e:
        return HttpResponse(f"Error de permisos: {e}", status=403)

    try:
        top_productos_data, titulo_reporte, fecha_inicio, fecha_fin = get_mas_vendidos_data(request)
    except Exception as e:
        return HttpResponse(f"Error al consultar la BD (Más Vendidos): {e}", status=500)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Top 10 Vendidos"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = titulo_reporte
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal="center")
        ws.append([]) 
        
        headers = ["Ranking", "ID Producto", "Nombre Producto", "Categoría", "Unidades Vendidas"]
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row_num = 4 
        for i, item in enumerate(top_productos_data):
            ws.append([
                i + 1,
                item['producto__id'],
                item['producto__nombre'],
                item['producto__categoria__nombre'],
                item['total_unidades']
            ])
            ws[f'E{row_num}'].number_format = '#,##0'
            row_num += 1

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'reporte_mas_vendidos_{fecha_inicio or "inicio"}_a_{fecha_fin or "fin"}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

    except Exception as e:
        traceback.print_exc()
        return HttpResponse("Error interno generando el Excel de Más Vendidos.", status=500)
